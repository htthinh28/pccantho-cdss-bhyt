#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Đọc Excel Luật PTTT/DVKT (sheet Danh_Sach_Luat) và sinh ma_nguon/tien_ich/du_lieu_luat_pttt_muc11.jsx

Cột bắt buộc: TRANG_THAI, MA_LUAT, TEN_QUY_TAC, DIEU_KIEN, CANH_BAO, GHI_CHU, NGUON_DU_LIEU

Chạy:
  python scripts/build_du_lieu_luat_pttt_muc11_from_excel.py
  python scripts/build_du_lieu_luat_pttt_muc11_from_excel.py --input "C:\\path\\DuLieu_LUAT_PTTT rule.xlsx"
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import date
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Cần cài: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUT = REPO_ROOT / "ma_nguon" / "tien_ich" / "du_lieu_luat_pttt_muc11.jsx"
DEFAULT_INPUT = Path(r"c:\Users\admin\Downloads\DuLieu_LUAT_PTTT rule.xlsx")

COLUMNS = [
    "TRANG_THAI",
    "MA_LUAT",
    "TEN_QUY_TAC",
    "DIEU_KIEN",
    "CANH_BAO",
    "GHI_CHU",
    "NGUON_DU_LIEU",
]


def cell_str(v) -> str:
    if v is None:
        return ""
    s = str(v).replace("\r\n", "\n").replace("\r", "\n").strip()
    return s


def validate_ma_luat(ma: str, row_idx: int) -> None:
    if not ma or ma.upper() in ("N/A", "NONE"):
        raise ValueError(f"Dòng {row_idx}: MA_LUAT trống hoặc không hợp lệ")
    if not re.match(r"^[A-Za-z0-9_.\-]+$", ma):
        raise ValueError(f"Dòng {row_idx}: MA_LUAT có ký tự lạ: {ma!r}")


def load_rows(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    if "Danh_Sach_Luat" not in sheet_names:
        wb.close()
        raise ValueError(f"Không thấy sheet Danh_Sach_Luat trong {path}. Có: {sheet_names}")
    ws = wb["Danh_Sach_Luat"]
    rows_iter = ws.iter_rows(values_only=True)
    header = [cell_str(x).upper() for x in next(rows_iter)]
    expected = [c.upper() for c in COLUMNS]
    if header[: len(expected)] != expected:
        raise ValueError(
            f"Header không đúng thứ tự.\nKỳ vọng: {expected}\nThực tế: {header[: len(expected)]}"
        )
    out: list[dict] = []
    for i, row in enumerate(rows_iter, start=2):
        if not row or all((x is None or str(x).strip() == "") for x in row):
            continue
        obj = {}
        for j, col in enumerate(COLUMNS):
            val = row[j] if j < len(row) else None
            obj[col] = cell_str(val)
        if not any(obj.values()):
            continue
        validate_ma_luat(obj["MA_LUAT"], i)
        tt = obj["TRANG_THAI"].upper() or "ON"
        if tt not in ("ON", "OFF"):
            raise ValueError(f"Dòng {i}: TRANG_THAI phải ON hoặc OFF, có: {obj['TRANG_THAI']!r}")
        obj["TRANG_THAI"] = tt
        if not obj["NGUON_DU_LIEU"]:
            obj["NGUON_DU_LIEU"] = path.name
        out.append(obj)
    wb.close()
    return out


def write_jsx(out_path: Path, rows: list[dict], source_name: str) -> None:
    stem = re.sub(r"[^A-Za-z0-9._\-]+", "_", Path(source_name).stem).strip("_") or "PTTT"
    version = f"{date.today().isoformat()}_muc11_pttt_{stem}"
    lines = [
        f"/** Auto-generated from {Path(source_name).name} — không sửa tay; chạy lại script build. */",
        f"export const PHIEN_BAN_SEED_LUAT_PTTT_MUC11 = {json.dumps(version, ensure_ascii=False)};",
        "export const COT_SEED_LUAT_PTTT_MUC11 = "
        + json.dumps(COLUMNS, ensure_ascii=False)
        + ";",
        "export const DU_LIEU_SEED_LUAT_PTTT_MUC11 = [",
    ]
    for obj in rows:
        lines.append("  " + json.dumps(obj, ensure_ascii=False) + ",")
    lines.append("];")
    lines.append("")
    text = "\n".join(lines)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(text, encoding="utf-8")
    print(f"Đã ghi {len(rows)} quy tắc → {out_path}")
    print(f"PHIEN_BAN_SEED_LUAT_PTTT_MUC11 = {version}")


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", type=Path, default=DEFAULT_INPUT, help="File .xlsx")
    ap.add_argument("--output", type=Path, default=DEFAULT_OUT, help="File .jsx đích")
    args = ap.parse_args()
    if not args.input.exists():
        print(f"Không tìm thấy file: {args.input}", file=sys.stderr)
        sys.exit(1)
    rows = load_rows(args.input)
    if len(rows) == 0:
        print("Không có dòng dữ liệu hợp lệ.", file=sys.stderr)
        sys.exit(1)
    write_jsx(args.output, rows, str(args.input))


if __name__ == "__main__":
    main()
